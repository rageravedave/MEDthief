"""
MEDthief — Desktop App für macOS
Entwickelt für MEDWING GmbH
"""

import sys
import os
import re
import subprocess
import threading
import webbrowser

try:
    import openpyxl
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QLineEdit, QSlider, QScrollArea, QFrame,
    QFileDialog, QProgressBar, QSizePolicy, QComboBox,
    QTextEdit, QMessageBox, QTabWidget, QInputDialog, QDialog,
    QToolButton, QMenu, QWidgetAction, QCheckBox,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QObject, QTimer, QUrl
from PyQt6.QtGui import QColor, QPalette, QCursor, QPixmap, QDesktopServices

from cv_parser import CVParser
from job_searcher import JobSearcher, compute_match_score
from profile_manager import ProfileManager

# ── Modern Dark Palette – readable, low-glare (shadcn / zinc-inspired) ──────
# Akzentfarben sind GEDÄMPFT, Text hat echten hohen Kontrast.
C_PRIMARY   = "#2DD4BF"   # teal-400  (accent, buttons)
C_PRIMARY_D = "#0F766E"   # teal-700  (pressed)
C_PRIMARY_L = "#99F6E4"   # teal-200  (soft highlight/links)

# Neutrale Zinc-Basis – kein grün, kein blau-stich
C_BG        = "#0A0A0B"   # near-black
C_SIDEBAR   = "#131316"
C_CARD      = "#17171A"   # zinc-900-ish
C_CARD_H    = "#1F1F23"
C_DIVIDER   = "#27272A"   # zinc-800
C_BORDER    = "#3F3F46"   # zinc-700

# Text: hohe Kontraste für echte Lesbarkeit
C_TEXT      = "#FAFAFA"   # near-white
C_SUB       = "#D4D4D8"   # zinc-300 (sekundärer Text – gut lesbar)
C_MUTED     = "#71717A"   # zinc-500 (labels, tertiary)
C_CONTACT   = "#5EEAD4"   # teal-300

C_WARN_FG   = "#FCD34D"   # amber-300 (gedämpftes Gelb statt knalliges Orange)
C_WARN_BG   = "#1C1403"

def _resource_path(relative: str) -> str:
    """Resolve resource path – works both in dev and PyInstaller bundle."""
    if getattr(sys, '_MEIPASS', None):
        return os.path.join(sys._MEIPASS, relative)
    return os.path.join(os.path.dirname(__file__), relative)

LOGO_PATH = _resource_path("medthief-logo.png")

# Status-Farben für CRM – pastellig, augenschonend
STATUS_CFG = {
    "Offen":        ("#A1A1AA", "#18181B"),   # zinc-400
    "Kontaktiert":  ("#7DD3FC", "#0C1B2A"),   # sky-300
    "In Gespräch":  ("#5EEAD4", "#0A1E1C"),   # teal-300
    "Abgesagt":     ("#FDA4AF", "#2A0B13"),   # rose-300
    "Vermittelt":   ("#86EFAC", "#0B2015"),   # green-300
}


def _mac_notify(title: str, body: str):
    """Sendet macOS-Desktop-Benachrichtigung via osascript."""
    try:
        safe_title = title.replace('"', '\\"')
        safe_body  = body.replace('"', '\\"')
        subprocess.Popen(
            ["osascript", "-e",
             f'display notification "{safe_body}" with title "{safe_title}"'],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
    except Exception:
        pass


STYLESHEET = f"""
QMainWindow, QWidget {{
    background-color: {C_BG};
    color: {C_TEXT};
    font-family: -apple-system, "SF Pro Text", "Helvetica Neue", Arial, sans-serif;
    font-size: 13px;
}}
QTabWidget::pane {{
    border: none;
    background-color: {C_BG};
}}
QTabBar::tab {{
    background-color: {C_SIDEBAR};
    color: {C_SUB};
    border: none;
    border-bottom: 2px solid transparent;
    padding: 10px 20px;
    font-size: 12px;
    font-weight: 600;
    min-width: 140px;
}}
QTabBar::tab:selected {{
    color: {C_TEXT};
    border-bottom: 2px solid {C_PRIMARY};
    background-color: {C_BG};
}}
QTabBar::tab:hover:!selected {{
    color: {C_TEXT};
    background-color: {C_CARD};
}}
QScrollArea {{
    border: none;
    background-color: transparent;
}}
QScrollBar:vertical {{
    background: transparent;
    width: 6px;
    margin: 0;
}}
QScrollBar::handle:vertical {{
    background: {C_BORDER};
    border-radius: 3px;
    min-height: 24px;
}}
QScrollBar::handle:vertical:hover {{
    background: {C_PRIMARY_D};
}}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
    height: 0px;
}}
QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
    background: transparent;
}}
QLineEdit {{
    background-color: #1D3430;
    color: {C_TEXT};
    border: 1px solid #2F3A48;
    border-radius: 10px;
    padding: 9px 14px;
    font-size: 13px;
    selection-background-color: {C_PRIMARY};
}}
QLineEdit:focus {{
    border: 1.5px solid {C_PRIMARY};
    background-color: #223C38;
}}
QLineEdit::placeholder {{
    color: #AEDBD5;
}}
QPushButton {{
    border-radius: 8px;
    font-weight: 600;
    font-size: 13px;
    padding: 9px 16px;
    border: none;
    outline: none;
}}
QPushButton:disabled {{
    opacity: 0.4;
}}
QSlider::groove:horizontal {{
    height: 3px;
    background: {C_BORDER};
    border-radius: 2px;
}}
QSlider::handle:horizontal {{
    background: white;
    border: 2px solid {C_PRIMARY};
    width: 14px;
    height: 14px;
    margin: -6px 0;
    border-radius: 7px;
}}
QSlider::sub-page:horizontal {{
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 {C_PRIMARY_D}, stop:1 {C_PRIMARY});
    border-radius: 2px;
}}
QProgressBar {{
    background-color: {C_BORDER};
    border-radius: 2px;
    height: 3px;
    text-align: center;
    border: none;
}}
QProgressBar::chunk {{
    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
        stop:0 {C_PRIMARY_D}, stop:1 {C_PRIMARY_L});
    border-radius: 2px;
}}
QComboBox {{
    background-color: #1D3430;
    color: {C_TEXT};
    border: 1px solid #2F3A48;
    border-radius: 10px;
    padding: 9px 14px;
    font-size: 13px;
    min-height: 20px;
}}
QComboBox:focus {{
    border: 1.5px solid {C_PRIMARY};
    background-color: #223C38;
}}
QComboBox QLineEdit {{
    background-color: transparent;
    color: {C_TEXT};
    border: none;
    padding: 0;
    font-size: 13px;
}}
QComboBox QLineEdit::placeholder {{
    color: #AEDBD5;
}}
QComboBox::drop-down {{
    border: none;
    width: 28px;
}}
QComboBox::down-arrow {{
    image: none;
    border-left: 4px solid transparent;
    border-right: 4px solid transparent;
    border-top: 5px solid {C_SUB};
    width: 0;
    height: 0;
    margin-right: 10px;
}}
QComboBox QAbstractItemView {{
    background-color: {C_CARD_H};
    color: {C_TEXT};
    border: 1px solid {C_BORDER};
    border-radius: 8px;
    selection-background-color: {C_PRIMARY};
    selection-color: white;
    padding: 4px;
    outline: none;
}}
QTextEdit {{
    background-color: {C_CARD};
    color: {C_TEXT};
    border: 1px solid {C_BORDER};
    border-radius: 10px;
    padding: 8px 12px;
    font-size: 12px;
}}
QTextEdit:focus {{
    border: 1.5px solid {C_PRIMARY};
}}
QToolTip {{
    background-color: {C_CARD_H};
    color: {C_TEXT};
    border: 1px solid {C_BORDER};
    border-radius: 6px;
    padding: 4px 8px;
    font-size: 11px;
}}
"""


# ── Email-Verify Worker ───────────────────────────────────────────────────────
class EmailVerifyWorker(QObject):
    finished = pyqtSignal(str)   # 'valid' | 'invalid' | 'unknown'

    def __init__(self, email: str):
        super().__init__()
        self._email = email

    def run(self):
        try:
            from smtp_verifier import verify_email
            result = verify_email(self._email)
        except Exception:
            result = "unknown"
        self.finished.emit(result)


# ── Worker-Thread ─────────────────────────────────────────────────────────────
class SearchWorker(QObject):
    progress = pyqtSignal(str)
    finished = pyqtSignal(list)
    error    = pyqtSignal(str)

    def __init__(self, searcher: JobSearcher, job_title, address, department,
                 einrichtung, radius, arbeitszeit="", schicht=""):
        super().__init__()
        self.searcher    = searcher
        self.job_title   = job_title
        self.address     = address
        self.department  = department
        self.einrichtung = einrichtung
        self.radius      = radius
        self.arbeitszeit = arbeitszeit
        self.schicht     = schicht

    def run(self):
        try:
            jobs = self.searcher.search(
                job_title=self.job_title,
                address=self.address,
                department=self.department,
                einrichtung=self.einrichtung,
                radius=self.radius,
                arbeitszeit=self.arbeitszeit,
                schicht=self.schicht,
                progress_cb=lambda msg: self.progress.emit(msg),
            )
            self.finished.emit(jobs)
        except Exception as e:
            self.error.emit(str(e))


# ── Kontakt-Template Dialog ───────────────────────────────────────────────────
class TemplateDialog(QDialog):
    def __init__(self, job: dict, candidate: dict, parent=None):
        super().__init__(parent)
        self._candidate = candidate
        self.setWindowTitle("Akquise-Anschreiben")
        self.setMinimumWidth(620)
        self.setMinimumHeight(520)
        self.setStyleSheet(f"background-color: {C_CARD}; color: {C_TEXT};")

        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)

        # ── Job-Daten ─────────────────────────────────────────────────────
        company = job.get("company", "")
        contact = job.get("contact_name", "")
        email   = job.get("contact_email", "")
        region  = job.get("location", "").split(",")[0].strip() or ""

        # ── Kandidaten-Daten ──────────────────────────────────────────────
        c_title   = candidate.get("job_title", "")
        c_fach    = candidate.get("fachabteilungen", "")
        c_start   = candidate.get("verfuegbar_ab", "")
        c_einr    = candidate.get("einrichtung", "")
        c_zeit    = candidate.get("arbeitszeit", "")
        c_schicht = candidate.get("schichten", "")

        # Ist die Einrichtung ein Seniorenheim? → Fachabt. weglassen
        _seniorenheim_keywords = (
            "altenheim", "pflegeheim", "seniorenheim", "altenpflege",
            "seniorenpflege", "seniorenzentrum", "stationäre pflege"
        )
        is_seniorenheim = any(
            kw in (c_einr or "").lower() or kw in (company or "").lower()
            for kw in _seniorenheim_keywords
        )

        # Anrede
        if contact:
            greeting = f"Sehr geehrte/r {contact},"
        else:
            greeting = "Sehr geehrte Damen und Herren,"

        # Kandidaten-Profil-Absatz dynamisch zusammenbauen
        profil_parts = []
        if c_title:
            profil_parts.append(f"examinierte Profile als {c_title}")
        if c_fach and not is_seniorenheim:
            profil_parts.append(f"z.\u202fB. für den Bereich {c_fach}")
        elif not is_seniorenheim:
            profil_parts.append("z.\u202fB. für den Bereich Intensivpflege oder Geriatrie")
        if c_start:
            profil_parts.append(f"verfügbar ab {c_start}")
        elif c_zeit:
            profil_parts.append(f"{c_zeit}")
        profil_hint = ", ".join(profil_parts) if profil_parts else \
            "examinierte Profile im Portfolio (z.\u202fB. für den Bereich Intensivpflege oder Geriatrie)"

        region_hint = f"in {region}" if region else "in Ihrer Region"

        # Betreff
        subj_title = c_title or "Pflegefachkraft"
        betreff = f"Passende/r {subj_title} für Ihre ausgeschriebene Stelle"

        # Kandidaten-Steckbrief (Bullet-Liste)
        profil_bullets = []
        if c_title:
            profil_bullets.append(f"Qualifikation: {c_title}")
        if c_fach and not is_seniorenheim:
            fach_parts = [f.strip() for f in re.split(r'[,;]', c_fach) if f.strip()]
            if len(fach_parts) == 1:
                profil_bullets.append(f"Erfahrung in: {fach_parts[0]}")
            elif fach_parts:
                profil_bullets.append("Erfahrung in: " + ", ".join(fach_parts))
        if c_start:
            profil_bullets.append(f"Verfügbar ab: {c_start}")
        if c_zeit:
            profil_bullets.append(f"Arbeitszeit: {c_zeit}")
        profil_block = "\n".join(f"  \u2022 {b}" for b in profil_bullets) if profil_bullets else ""

        template = (
            f"{betreff}\n\n"
            f"{greeting}\n\n"
            f"mein Name ist David Böser, ich bin Karriereberater bei MEDWING \u2013 "
            f"spezialisiert auf die Direktvermittlung von Pflegefachkräften und "
            f"medizinischem Personal in die Festanstellung. Deutschlandweit "
            f"unterstützen wir Einrichtungen dabei, offene Stellen schnell und "
            f"passgenau zu besetzen.\n\n"
            f"Ich schreibe Ihnen, weil ich aktuell eine/n examinierte/n "
            f"{subj_title} betreue, die/der sich gezielt auf Ihre "
            f"ausgeschriebene Stelle bewirbt und die Anforderungen "
            f"sehr gut erfüllt.\n\n"
        )
        if profil_block:
            template += f"Kurzes Kandidatenprofil:\n{profil_block}\n\n"
        template += (
            f"Was eine Zusammenarbeit mit MEDWING für Sie bedeutet:\n\n"
            f"  \u2713 Niedrige Gebühr: Unsere Vermittlungspauschale liegt "
            f"deutlich unter dem Branchenschnitt und häufig auch unter "
            f"Ihren eigenen Recruiting-Kosten.\n\n"
            f"  \u2713 Abgesichertes Risiko: Bei Nichtantritt erhalten Sie die "
            f"Gebühr vollständig zurück. Endet das Arbeitsverhältnis während "
            f"der Probezeit, greift unsere gestaffelte Rückvergütung.\n\n"
            f"  \u2713 Kein Aufwand: Wir übernehmen Vorqualifizierung, "
            f"Dokumentenprüfung und Koordination. "
            f"Sie führen nur noch das Gespräch.\n\n"
            f"Ich sende Ihnen das anonymisierte Kurzprofil gerne zu. "
            f"Falls Sie zunächst mehr über unsere Arbeitsweise oder die "
            f"Konditionen einer Kooperationsvereinbarung erfahren möchten, "
            f"stehe ich selbstverständlich auch dafür zur Verfügung. "
            f"Eine kurze Rückmeldung genügt, ich melde mich dann "
            f"innerhalb von 24 Stunden.\n\n"
            f"Beste Grüße\n"
            f"David Böser\n"
            f"MEDWING GmbH \u00b7 Recruiting"
        )

        hdr = QLabel("Akquise-Vorlage anpassen und kopieren:")
        hdr.setStyleSheet(f"color: {C_SUB}; font-size: 11px;")
        layout.addWidget(hdr)

        self.text_edit = QTextEdit()
        self.text_edit.setPlainText(template)
        self.text_edit.setMinimumHeight(340)
        self.text_edit.setStyleSheet(f"""
            QTextEdit {{
                background: {C_BG};
                color: {C_TEXT};
                border: 1px solid {C_BORDER};
                border-radius: 8px;
                padding: 12px;
                font-size: 13px;
                line-height: 1.5;
            }}
        """)
        layout.addWidget(self.text_edit)

        btn_row = QHBoxLayout()
        btn_copy = QPushButton("📋  In Zwischenablage kopieren")
        btn_copy.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                    stop:0 #45C9B4, stop:1 #2A9880);
                color: white; border-radius: 8px;
                font-size: 12px; font-weight: 700; padding: 8px 18px;
            }}
            QPushButton:hover {{ background: {C_PRIMARY}; }}
        """)
        btn_copy.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_copy.clicked.connect(
            lambda: QApplication.clipboard().setText(self.text_edit.toPlainText())
        )

        btn_mailto = QPushButton("📧  In Mail öffnen")
        btn_mailto.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                    stop:0 #6366F1, stop:1 #4338CA);
                color: white; border-radius: 8px;
                font-size: 12px; font-weight: 700; padding: 8px 18px;
            }}
            QPushButton:hover {{ background: #818CF8; }}
        """)
        btn_mailto.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self._email = email
        self._betreff = betreff
        if email:
            btn_mailto.clicked.connect(self._open_gmail)
        else:
            btn_mailto.setEnabled(False)
            btn_mailto.setToolTip("Keine E-Mail-Adresse vorhanden")

        btn_close = QPushButton("Schließen")
        btn_close.setStyleSheet(f"""
            QPushButton {{
                background: transparent; color: {C_SUB};
                border: 1px solid {C_BORDER}; border-radius: 8px;
                font-size: 12px; padding: 8px 16px;
            }}
            QPushButton:hover {{ color: {C_TEXT}; border-color: {C_SUB}; }}
        """)
        btn_close.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_close.clicked.connect(self.accept)
        btn_row.addWidget(btn_copy)
        btn_row.addWidget(btn_mailto)
        btn_row.addWidget(btn_close)
        layout.addLayout(btn_row)

        # Hinweis unter den Buttons
        hint = QLabel("Der anonymisierte Kandidaten-PDF wird beim Mail-Versand automatisch im Finder geöffnet.")
        hint.setWordWrap(True)
        hint.setStyleSheet(f"color: {C_MUTED}; font-size: 10px; padding-top: 4px;")
        layout.addWidget(hint)

    def _build_anon_pdf(self) -> str:
        """Erzeugt ein anonymisiertes Kandidatenprofil als PDF. Gibt Pfad zurück."""
        from fpdf import FPDF
        import tempfile

        c = self._candidate

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=20)
        pdf.add_page()

        # Header
        pdf.set_fill_color(15, 118, 110)  # teal-700
        pdf.rect(0, 0, 210, 38, "F")
        pdf.set_font("Helvetica", "B", 18)
        pdf.set_text_color(255, 255, 255)
        pdf.set_y(10)
        pdf.cell(0, 10, "MEDWING - Kandidatenprofil (anonymisiert)", align="C", new_x="LMARGIN", new_y="NEXT")

        pdf.ln(20)
        pdf.set_text_color(30, 30, 30)

        def _section(title):
            pdf.set_font("Helvetica", "B", 12)
            pdf.set_fill_color(230, 245, 243)
            pdf.cell(0, 9, f"  {title}", fill=True, new_x="LMARGIN", new_y="NEXT")
            pdf.ln(3)
            pdf.set_font("Helvetica", "", 11)

        def _row(label, value):
            if not value:
                return
            pdf.set_font("Helvetica", "B", 11)
            pdf.cell(55, 7, f"{label}:", new_x="END")
            pdf.set_font("Helvetica", "", 11)
            pdf.cell(0, 7, value, new_x="LMARGIN", new_y="NEXT")

        _section("Qualifikation")
        _row("Beruf", c.get("job_title", ""))
        _row("Einrichtungsart", c.get("einrichtung", ""))
        fach = c.get("fachabteilungen", "")
        if fach:
            _row("Fachbereiche", fach)
        pdf.ln(4)

        _section("Verfügbarkeit")
        _row("Verfügbar ab", c.get("verfuegbar_ab", "") or "auf Anfrage")
        _row("Arbeitszeit", c.get("arbeitszeit", "") or "flexibel")
        schicht = c.get("schichten", "")
        if schicht:
            _row("Schichtbereitschaft", schicht)
        _row("Region", c.get("wohnort", ""))
        pdf.ln(4)

        _section("Hinweis")
        pdf.set_font("Helvetica", "I", 10)
        pdf.set_text_color(100, 100, 100)
        pdf.multi_cell(0, 6,
            "Dieses Profil wurde zum Schutz der Kandidatendaten anonymisiert. "
            "Persönliche Daten (Name, Adresse, Kontakt) werden erst nach "
            "einer Kooperationsvereinbarung mit MEDWING übermittelt."
        )

        # Footer
        pdf.set_y(-25)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(150, 150, 150)
        pdf.cell(0, 5, "MEDWING GmbH | Vertraulich", align="C")

        path = os.path.join(tempfile.gettempdir(), "MEDWING_Kandidatenprofil_anonym.pdf")
        pdf.output(path)
        return path

    def _open_gmail(self):
        """Öffnet Gmail Compose + anonymisierten PDF im Finder zum Anhängen."""
        import urllib.parse as _up

        # 1. Anonymisierten PDF erzeugen und im Finder markieren
        try:
            pdf_path = self._build_anon_pdf()
            subprocess.Popen(["open", "-R", pdf_path])
        except Exception as e:
            print(f"[Anon-PDF] Fehler: {e}")

        # 2. Gmail öffnen
        body = self.text_edit.toPlainText()
        subj = self._betreff
        gmail_url = (
            "https://mail.google.com/mail/?view=cm&fs=1"
            f"&to={_up.quote(self._email)}"
            f"&su={_up.quote(subj)}"
            f"&body={_up.quote(body)}"
        )
        QDesktopServices.openUrl(QUrl(gmail_url))


# ── Multi-Select Dropdown (Fachabteilungen) ─────────────────────────────────
_CB_STYLE = """
    QCheckBox {
        color: #F4F6F8;
        padding: 5px 10px;
        font-size: 12px;
        spacing: 8px;
        background: transparent;
    }
    QCheckBox:hover { background: #1F2937; border-radius: 4px; }
    QCheckBox::indicator { width: 14px; height: 14px; }
    QCheckBox::indicator:unchecked {
        border: 1.5px solid #64748B;
        border-radius: 3px;
        background: #0F172A;
    }
    QCheckBox::indicator:checked {
        border: 1.5px solid #38A694;
        border-radius: 3px;
        background: #38A694;
    }
"""


class MultiSelectButton(QToolButton):
    """Button mit Popup-Menü: Suchfeld oben + alphabetische Checkbox-Liste.
    Unbekannte Eingaben können als neue Custom-Fachbereiche hinzugefügt werden."""

    def __init__(self, items: list, placeholder: str = "Auswählen …", parent=None):
        super().__init__(parent)
        self._placeholder = placeholder
        self._checks: dict = {}  # label → QCheckBox
        self._order: list = []   # sortierte Reihenfolge der Labels

        self.setText(placeholder)
        self.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.setMinimumHeight(36)
        self.setStyleSheet(f"""
            QToolButton {{
                background-color: #1D3430;
                color: #F4F6F8;
                border: 1px solid #2F3A48;
                border-radius: 10px;
                padding: 8px 14px;
                font-size: 13px;
                text-align: left;
            }}
            QToolButton:hover {{ border: 1.5px solid #38A694; }}
            QToolButton::menu-indicator {{
                subcontrol-position: right center;
                subcontrol-origin: padding;
                right: 10px;
            }}
        """)

        # ── Popup-Container: Suchfeld + Scroll-Liste ─────────────────────────
        self._menu = QMenu(self)
        self._menu.setStyleSheet("""
            QMenu { background: #141A22; border: 1px solid #2F3A48; padding: 0; }
        """)

        container = QWidget()
        container.setFixedWidth(280)
        cl = QVBoxLayout(container)
        cl.setContentsMargins(8, 8, 8, 8)
        cl.setSpacing(6)

        # Suchfeld
        self._search = QLineEdit()
        self._search.setPlaceholderText("Suchen oder neues Fach eingeben …")
        self._search.setStyleSheet("""
            QLineEdit {
                background: #0F172A;
                color: #F4F6F8;
                border: 1px solid #2F3A48;
                border-radius: 6px;
                padding: 6px 10px;
                font-size: 12px;
            }
            QLineEdit:focus { border: 1.5px solid #38A694; }
        """)
        self._search.textChanged.connect(self._on_search_changed)
        self._search.returnPressed.connect(self._on_search_enter)
        cl.addWidget(self._search)

        # "+ Hinzufügen"-Button (taucht nur bei Custom-Text auf)
        self._add_btn = QPushButton("")
        self._add_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self._add_btn.setStyleSheet("""
            QPushButton {
                background: #164E47;
                color: #99F6E4;
                border: 1px solid #38A694;
                border-radius: 6px;
                padding: 6px 10px;
                font-size: 12px;
                text-align: left;
            }
            QPushButton:hover { background: #1D6B61; }
        """)
        self._add_btn.clicked.connect(self._on_search_enter)
        self._add_btn.hide()
        cl.addWidget(self._add_btn)

        # Scroll-Bereich mit allen Checkboxen
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFixedHeight(320)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet("""
            QScrollArea { background: transparent; border: none; }
            QScrollBar:vertical {
                background: #0F172A; width: 8px; border-radius: 4px;
            }
            QScrollBar::handle:vertical {
                background: #38A694; border-radius: 4px; min-height: 20px;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }
        """)
        list_widget = QWidget()
        list_widget.setStyleSheet("background: transparent;")
        self._list_layout = QVBoxLayout(list_widget)
        self._list_layout.setContentsMargins(0, 0, 0, 0)
        self._list_layout.setSpacing(1)
        scroll.setWidget(list_widget)
        cl.addWidget(scroll)

        # Footer: "Alle löschen"
        clear_btn = QPushButton("Alle löschen")
        clear_btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        clear_btn.setStyleSheet("""
            QPushButton {
                background: transparent;
                color: #94A3B8;
                border: none;
                padding: 4px;
                font-size: 11px;
            }
            QPushButton:hover { color: #F4F6F8; }
        """)
        clear_btn.clicked.connect(self.clear_selection)
        cl.addWidget(clear_btn)

        wa = QWidgetAction(self._menu)
        wa.setDefaultWidget(container)
        self._menu.addAction(wa)
        self.setMenu(self._menu)

        # Initial items alphabetisch einfügen
        for item in sorted(items, key=lambda s: s.lower()):
            self._add_checkbox(item)

    # ── intern ────────────────────────────────────────────────────────────
    def _add_checkbox(self, label: str):
        if label in self._checks:
            return
        cb = QCheckBox(label)
        cb.setStyleSheet(_CB_STYLE)
        cb.toggled.connect(self._refresh_label)
        # Sortiert einfügen
        ll = label.lower()
        insert_at = 0
        for i, existing in enumerate(self._order):
            if existing.lower() > ll:
                break
            insert_at = i + 1
        self._order.insert(insert_at, label)
        self._list_layout.insertWidget(insert_at, cb)
        self._checks[label] = cb

    def _on_search_changed(self, text: str):
        q = text.strip().lower()
        visible = 0
        exact_match = False
        for label, cb in self._checks.items():
            if not q or q in label.lower():
                cb.show()
                visible += 1
            else:
                cb.hide()
            if label.lower() == q:
                exact_match = True
        # Add-Button nur zeigen, wenn Suchtext vorhanden & kein Exact-Match
        if q and not exact_match and len(q) >= 2:
            self._add_btn.setText(f"+  '{text.strip()}' hinzufügen")
            self._add_btn.show()
        else:
            self._add_btn.hide()

    def _on_search_enter(self):
        text = self._search.text().strip()
        if not text or len(text) < 2:
            return
        # Gibt's schon? Dann anhaken statt neu anlegen
        for label, cb in self._checks.items():
            if label.lower() == text.lower():
                cb.setChecked(True)
                self._search.clear()
                self._on_search_changed("")
                return
        # Neu anlegen + anhaken
        self._add_checkbox(text)
        self._checks[text].setChecked(True)
        self._search.clear()
        self._on_search_changed("")

    def _refresh_label(self):
        sel = self.selected()
        if not sel:
            self.setText(self._placeholder)
        elif len(sel) == 1:
            self.setText(sel[0])
        else:
            self.setText(f"{sel[0]} +{len(sel)-1}")

    # ── öffentliche API ───────────────────────────────────────────────────
    def selected(self) -> list:
        return [label for label in self._order if self._checks[label].isChecked()]

    def text_value(self) -> str:
        return ", ".join(self.selected())

    def clear_selection(self):
        for cb in self._checks.values():
            cb.blockSignals(True)
            cb.setChecked(False)
            cb.blockSignals(False)
        self._refresh_label()

    def set_from_text(self, text: str):
        """Markiert Items die im Text vorkommen (kommagetrennt, fuzzy lowercase).
        Unbekannte Werte werden als Custom-Items hinzugefügt."""
        self.clear_selection()
        if not text:
            return
        parts = [p.strip() for p in re.split(r'[,;/]|\bund\b', text) if p.strip()]
        for part in parts:
            pl = part.lower()
            matched_label = None
            # Exakter Match zuerst
            for label in self._order:
                if label.lower() == pl:
                    matched_label = label
                    break
            # Fuzzy Substring-Match
            if not matched_label:
                for label in self._order:
                    ll = label.lower()
                    if pl in ll or ll in pl:
                        matched_label = label
                        break
            # Nichts gefunden → als Custom-Item anlegen
            if not matched_label:
                self._add_checkbox(part)
                matched_label = part
            cb = self._checks.get(matched_label)
            if cb:
                cb.blockSignals(True)
                cb.setChecked(True)
                cb.blockSignals(False)
        self._refresh_label()


# ── Job-Karte ─────────────────────────────────────────────────────────────────
class JobCard(QFrame):
    SOURCE_COLORS = {
        # Pastell-Palette (Tailwind *-300) – augenschonend auf Dark-BG
        "Pflegia":         "#F9A8D4",  # pink-300
        "Indeed":          "#93C5FD",  # blue-300
        "LinkedIn":        "#7DD3FC",  # sky-300
        "Pflegejobs":      "#FDBA74",  # orange-300
        "Medi-Karriere":   "#FDA4AF",  # rose-300
        "Kliniken.de":     "#67E8F9",  # cyan-300
        "Gesundheit.jobs": "#86EFAC",  # green-300
    }
    SOURCE_DISPLAY = {
        "Pflegia":         "Pflegia.de",
        "Indeed":          "Indeed",
        "LinkedIn":        "LinkedIn",
        "Pflegejobs":      "Pflegejobs.de",
        "Medi-Karriere":   "Medi-Karriere",
        "Kliniken.de":     "Kliniken.de",
        "Gesundheit.jobs": "Gesundheit.jobs",
    }

    def __init__(self, job: dict, ctx: dict, parent=None):
        """
        ctx keys:
          job_status       {url -> status_str}
          job_notes        {url -> note_str}
          seen_urls        set of previously seen urls
          on_status_change callable(url, status)
          on_notes_change  callable(url, note)
        """
        super().__init__(parent)
        self.job  = job
        self.ctx  = ctx
        self._verify_thread = None
        self._verify_worker = None
        self._color = self.SOURCE_COLORS.get(job.get("source", ""), "#4A6A65")
        # Nur so hoch wie der Inhalt – verhindert, dass eine einzelne Karte
        # die komplette ScrollArea vertikal ausfüllt.
        self.setSizePolicy(
            QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed
        )
        self._build()

    def _build(self):
        job    = self.job
        ctx    = self.ctx
        source = job.get("source", "")
        color  = self._color
        label  = self.SOURCE_DISPLAY.get(source, source)
        url    = job.get("url", "")

        self.setStyleSheet(f"""
            JobCard {{
                background-color: {C_CARD};
                border-radius: 14px;
                border-left: 4px solid {color};
            }}
        """)
        self.setCursor(QCursor(Qt.CursorShape.ArrowCursor))

        outer = QVBoxLayout(self)
        outer.setContentsMargins(18, 14, 18, 14)
        outer.setSpacing(0)

        # ── Row 1: Badge + NEU + Titel + Score ────────────────────────────
        row1 = QHBoxLayout()
        row1.setSpacing(8)

        badge = QLabel(label)
        badge.setSizePolicy(QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Fixed)
        badge.setStyleSheet(f"""
            background-color: #22{color[1:]}; color: {color};
            border: 1px solid #55{color[1:]}; border-radius: 10px;
            padding: 3px 10px; font-size: 10px; font-weight: 700;
        """)
        row1.addWidget(badge)

        # NEU-Badge wenn URL noch nicht gesehen
        if url and url not in ctx.get("seen_urls", set()):
            neu = QLabel("NEU")
            neu.setStyleSheet("""
                background-color: #2286EFAC; color: #86EFAC;
                border: 1px solid #4486EFAC; border-radius: 8px;
                padding: 2px 8px; font-size: 10px; font-weight: 800;
            """)
            row1.addWidget(neu)

        title_lbl = QLabel(job.get("title", "(Kein Titel)"))
        title_lbl.setStyleSheet(f"color: {C_TEXT}; font-size: 14px; font-weight: 700;")
        title_lbl.setWordWrap(True)
        row1.addWidget(title_lbl, stretch=1)

        # Match-Score Badge – pastellige, nicht-leuchtende Farben
        score = job.get("match_score")
        if score is not None:
            if score >= 80:
                sc_fg, sc_bg = "#86EFAC", "#0F1F17"   # green-300
            elif score >= 55:
                sc_fg, sc_bg = "#5EEAD4", "#0D1C1B"   # teal-300
            elif score >= 35:
                sc_fg, sc_bg = "#FCD34D", "#1C1403"   # amber-300
            else:
                sc_fg, sc_bg = "#A1A1AA", C_CARD       # zinc-400
            score_lbl = QLabel(f"{score}%")
            score_lbl.setToolTip(
                f"Match-Score: {score}%\n"
                "Berechnet aus: Titel, Einrichtungsart, Entfernung, Kontaktdaten."
            )
            score_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            score_lbl.setStyleSheet(f"""
                background-color: {sc_bg}; color: {sc_fg};
                border: 1px solid #33{sc_fg[1:]}; border-radius: 10px;
                padding: 3px 10px; font-size: 11px; font-weight: 800;
                min-width: 36px;
            """)
            row1.addWidget(score_lbl)

        outer.addLayout(row1)
        outer.addSpacing(8)

        # ── Row 2: Firma + Ort + Facility-Warning ─────────────────────────
        company  = job.get("company", "")
        location = job.get("location", "")
        dist     = job.get("distance_km")
        row2 = QHBoxLayout()
        row2.setSpacing(6)

        if company:
            co = QLabel(company)
            co.setStyleSheet(f"color: {C_TEXT}; font-size: 12px; font-weight: 500;")
            row2.addWidget(co)
        if company and location:
            sep = QLabel("·")
            sep.setStyleSheet(f"color: {C_DIVIDER}; font-size: 12px;")
            row2.addWidget(sep)
        if location:
            dist_str = f"  {dist:.0f} km" if dist is not None else ""
            loc = QLabel(f"📍 {location}{dist_str}")
            loc.setStyleSheet(f"color: {C_SUB}; font-size: 12px;")
            row2.addWidget(loc)

        row2.addStretch()

        if not job.get("facility_match", True):
            detected = job.get("detected_facility", "")
            if detected:
                warn = QLabel(f"⚠  {detected}")
                warn.setStyleSheet(f"""
                    color: {C_WARN_FG}; background-color: {C_WARN_BG};
                    border: 1px solid #44{C_WARN_FG[1:]}; border-radius: 8px;
                    padding: 2px 8px; font-size: 10px; font-weight: 600;
                """)
                row2.addWidget(warn)

        outer.addLayout(row2)

        # ── Divider ───────────────────────────────────────────────────────
        sep_line = QFrame()
        sep_line.setFrameShape(QFrame.Shape.HLine)
        sep_line.setFixedHeight(1)
        sep_line.setStyleSheet(f"background-color: {C_DIVIDER}; border: none; margin: 8px 0;")
        outer.addSpacing(8)
        outer.addWidget(sep_line)
        outer.addSpacing(8)

        # ── Row 3: Kontaktdaten ───────────────────────────────────────────
        self._build_contact_row(outer, job, color)

        # ── Row 4: CRM Status-Buttons ─────────────────────────────────────
        outer.addSpacing(10)
        self._build_status_row(outer, url, ctx)

        # ── Row 5: Toolbar (Anschreiben | Notiz | Zur Stelle) ─────────────
        outer.addSpacing(8)
        row5 = QHBoxLayout()
        row5.setSpacing(8)

        pub = job.get("published", "")
        if pub:
            pub_lbl = QLabel(pub)
            pub_lbl.setStyleSheet(f"color: {C_SUB}; font-size: 10px;")
            row5.addWidget(pub_lbl)
        row5.addStretch()

        btn_tpl = QPushButton("✉  Anschreiben")
        btn_tpl.setStyleSheet(self._toolbar_btn_style())
        btn_tpl.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_tpl.clicked.connect(
            lambda: TemplateDialog(self.job, self.ctx.get("candidate_info", {}), self).exec()
        )
        row5.addWidget(btn_tpl)

        existing_note = ctx.get("job_notes", {}).get(url, "")
        self._btn_note = QPushButton(
            "📝  Notiz  ▸" if existing_note else "📝  Notiz"
        )
        self._btn_note.setStyleSheet(self._toolbar_btn_style())
        self._btn_note.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self._btn_note.clicked.connect(lambda: self._toggle_notes(url, ctx))
        row5.addWidget(self._btn_note)

        if url:
            btn_link = QPushButton("Zur Stelle  →")
            btn_link.setStyleSheet(f"""
                QPushButton {{
                    background-color: #20{color[1:]}; color: {color};
                    font-size: 11px; font-weight: 600; padding: 5px 14px;
                    border-radius: 8px; border: 1px solid #44{color[1:]};
                }}
                QPushButton:hover {{
                    background-color: #38{color[1:]}; border-color: #88{color[1:]};
                }}
            """)
            btn_link.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            btn_link.clicked.connect(lambda _, u=url: webbrowser.open(u))
            row5.addWidget(btn_link)

        outer.addLayout(row5)

        # ── Notizen-Bereich (collapsible) ──────────────────────────────────
        self._notes_container = QWidget()
        self._notes_container.setVisible(bool(existing_note))
        nc_layout = QVBoxLayout(self._notes_container)
        nc_layout.setContentsMargins(0, 6, 0, 0)
        nc_layout.setSpacing(0)

        self._notes_edit = QTextEdit()
        self._notes_edit.setPlaceholderText("Notiz zu dieser Stelle …")
        self._notes_edit.setFixedHeight(68)
        self._notes_edit.setPlainText(existing_note)
        self._notes_edit.setStyleSheet(f"""
            QTextEdit {{
                background-color: {C_CARD_H}; color: {C_TEXT};
                border: 1px solid {C_BORDER}; border-radius: 8px;
                padding: 6px 10px; font-size: 12px;
            }}
            QTextEdit:focus {{ border: 1px solid {C_PRIMARY}; }}
        """)
        self._notes_edit.textChanged.connect(
            lambda: self._on_notes_changed(url, ctx)
        )
        nc_layout.addWidget(self._notes_edit)
        outer.addWidget(self._notes_container)

    # ── Contact row ───────────────────────────────────────────────────────
    def _build_contact_row(self, outer, job, color):
        row = QHBoxLayout()
        row.setSpacing(6)

        name  = job.get("contact_name", "")
        role  = job.get("contact_role", "")
        phone = job.get("contact_phone", "")
        email = job.get("contact_email", "")

        if name:
            name_str = f"👤  {name}"
            if role:
                name_str += f"  ·  {role}"
            nl = QLabel(name_str)
            nl.setStyleSheet(f"color: {C_CONTACT}; font-size: 11px; font-weight: 500;")
            row.addWidget(nl)
            if phone or email:
                row.addWidget(QLabel("  "))

        if phone:
            phone_clean = re.sub(r"[^\d+]", "", phone)
            btn_p = QPushButton(f"📞  {phone}")
            btn_p.setStyleSheet(self._contact_btn_style(C_PRIMARY))
            btn_p.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            btn_p.clicked.connect(
                lambda _, p=phone_clean: QDesktopServices.openUrl(QUrl(f"tel:{p}"))
            )
            row.addWidget(btn_p)
            row.addWidget(self._copy_btn(phone))

        if email:
            is_guessed = job.get("contact_email_guessed", False)
            btn_m = QPushButton(f"✉  {email}")
            btn_m.setStyleSheet(
                self._contact_btn_style(C_WARN_FG if is_guessed else C_PRIMARY)
            )
            if is_guessed:
                btn_m.setToolTip("⚠ Geschätzte Email — bitte vor dem Senden verifizieren!")
            btn_m.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            btn_m.clicked.connect(
                lambda _, m=email: QDesktopServices.openUrl(QUrl(f"mailto:{m}"))
            )
            row.addWidget(btn_m)
            row.addWidget(self._copy_btn(email))

            if is_guessed:
                self._guess_badge = QLabel("?")
                self._guess_badge.setToolTip(
                    "Email automatisch geschätzt — könnte falsch sein."
                )
                self._guess_badge.setStyleSheet(f"""
                    background-color: #33{C_WARN_FG[1:]}; color: {C_WARN_FG};
                    border: 1px solid #77{C_WARN_FG[1:]}; border-radius: 8px;
                    padding: 2px 7px; font-size: 10px; font-weight: 800;
                """)
                row.addWidget(self._guess_badge)

                self._btn_verify = QPushButton("Prüfen")
                self._btn_verify.setToolTip("SMTP-Prüfung — sendet keine E-Mail")
                self._btn_verify.setStyleSheet(f"""
                    QPushButton {{
                        background: transparent; color: {C_WARN_FG};
                        border: 1px solid #55{C_WARN_FG[1:]}; border-radius: 6px;
                        font-size: 10px; font-weight: 600; padding: 2px 8px;
                    }}
                    QPushButton:hover {{ background: #22{C_WARN_FG[1:]}; }}
                    QPushButton:disabled {{ opacity: 0.4; }}
                """)
                self._btn_verify.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
                self._btn_verify.clicked.connect(
                    lambda _, e=email: self._start_verify(e)
                )
                row.addWidget(self._btn_verify)

        if not name and not phone and not email:
            no = QLabel("Keine Kontaktdaten verfügbar")
            no.setStyleSheet(f"color: {C_SUB}; font-size: 11px; font-style: italic;")
            row.addWidget(no)

        row.addStretch()
        outer.addLayout(row)

    # ── Status row ────────────────────────────────────────────────────────
    def _build_status_row(self, outer, url, ctx):
        row = QHBoxLayout()
        row.setSpacing(4)

        lbl = QLabel("Status:")
        lbl.setStyleSheet(f"color: {C_SUB}; font-size: 10px; font-weight: 600;")
        row.addWidget(lbl)

        current = ctx.get("job_status", {}).get(url, "Offen")
        self._status_btns = {}
        for st, (fg, bg) in STATUS_CFG.items():
            btn = QPushButton(st)
            btn.setStyleSheet(self._status_btn_style(fg, bg, st == current))
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            btn.clicked.connect(lambda _, s=st: self._set_status(s, url, ctx))
            self._status_btns[st] = btn
            row.addWidget(btn)

        row.addStretch()
        outer.addLayout(row)

    # ── Callbacks ─────────────────────────────────────────────────────────
    def _set_status(self, new_status: str, url: str, ctx: dict):
        ctx.setdefault("job_status", {})[url] = new_status
        for st, btn in self._status_btns.items():
            fg, bg = STATUS_CFG[st]
            btn.setStyleSheet(self._status_btn_style(fg, bg, st == new_status))
        cb = ctx.get("on_status_change")
        if cb:
            cb(url, new_status)

    def _toggle_notes(self, url, ctx):
        visible = not self._notes_container.isVisible()
        self._notes_container.setVisible(visible)
        existing = ctx.get("job_notes", {}).get(url, "")
        if visible:
            self._btn_note.setText("📝  Notiz  ▾")
            self._notes_edit.setFocus()
        else:
            self._btn_note.setText("📝  Notiz  ▸" if existing else "📝  Notiz")

    def _on_notes_changed(self, url: str, ctx: dict):
        text = self._notes_edit.toPlainText()
        ctx.setdefault("job_notes", {})[url] = text
        cb = ctx.get("on_notes_change")
        if cb:
            cb(url, text)
        self._btn_note.setText("📝  Notiz  ▸" if text.strip() else "📝  Notiz")

    # ── Email-Verify ──────────────────────────────────────────────────────
    def _start_verify(self, email: str):
        self._btn_verify.setEnabled(False)
        self._btn_verify.setText("Prüft …")
        self._verify_thread = QThread()
        self._verify_worker = EmailVerifyWorker(email)
        self._verify_worker.moveToThread(self._verify_thread)
        self._verify_thread.started.connect(self._verify_worker.run)
        self._verify_worker.finished.connect(self._on_verify_done)
        self._verify_worker.finished.connect(self._verify_thread.quit)
        self._verify_thread.start()

    def _on_verify_done(self, result: str):
        if result == "valid":
            self._guess_badge.setText("✓")
            self._guess_badge.setToolTip("E-Mail existiert (SMTP bestätigt)")
            self._guess_badge.setStyleSheet("""
                background-color: #2286EFAC; color: #86EFAC;
                border: 1px solid #4486EFAC; border-radius: 8px;
                padding: 2px 7px; font-size: 10px; font-weight: 800;
            """)
            self._btn_verify.hide()
        elif result == "invalid":
            self._guess_badge.setText("✗")
            self._guess_badge.setToolTip("E-Mail existiert nicht (SMTP abgewiesen)")
            self._guess_badge.setStyleSheet("""
                background-color: #22FDA4AF; color: #FDA4AF;
                border: 1px solid #44FDA4AF; border-radius: 8px;
                padding: 2px 7px; font-size: 10px; font-weight: 800;
            """)
            self._btn_verify.hide()
        else:
            self._guess_badge.setToolTip("Konnte nicht verifiziert werden (Timeout)")
            self._btn_verify.setText("Erneut")
            self._btn_verify.setEnabled(True)

    # ── Style helpers ─────────────────────────────────────────────────────
    @staticmethod
    def _contact_btn_style(base: str) -> str:
        return f"""
            QPushButton {{
                background-color: #18{base[1:]}; color: {base};
                font-size: 11px; border: 1px solid #33{base[1:]};
                border-radius: 8px; padding: 3px 10px; font-weight: 500;
            }}
            QPushButton:hover {{
                background-color: #30{base[1:]}; border-color: #66{base[1:]};
            }}
        """

    @staticmethod
    def _copy_btn(value: str) -> QPushButton:
        btn = QPushButton("⎘")
        btn.setToolTip("Kopieren")
        btn.setStyleSheet(f"""
            QPushButton {{ background: transparent; color: {C_SUB};
                font-size: 12px; border: none; padding: 0 4px; }}
            QPushButton:hover {{ color: {C_TEXT}; }}
        """)
        btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn.clicked.connect(lambda _, v=value: QApplication.clipboard().setText(v))
        return btn

    @staticmethod
    def _status_btn_style(fg: str, bg: str, active: bool) -> str:
        if active:
            return f"""
                QPushButton {{
                    background-color: {bg}; color: {fg};
                    border: 1.5px solid #88{fg[1:]}; border-radius: 6px;
                    font-size: 10px; font-weight: 700; padding: 2px 8px;
                }}
            """
        return f"""
            QPushButton {{
                background-color: transparent; color: {C_SUB};
                border: 1px solid {C_BORDER}; border-radius: 6px;
                font-size: 10px; font-weight: 500; padding: 2px 8px;
            }}
            QPushButton:hover {{ color: {fg}; border-color: #55{fg[1:]}; }}
        """

    @staticmethod
    def _toolbar_btn_style() -> str:
        return f"""
            QPushButton {{
                background-color: transparent; color: {C_SUB};
                font-size: 11px; font-weight: 600; padding: 5px 12px;
                border-radius: 8px; border: 1px solid {C_BORDER};
            }}
            QPushButton:hover {{ color: {C_TEXT}; border-color: {C_SUB}; }}
        """

    def enterEvent(self, event):
        self.setStyleSheet(f"""
            JobCard {{
                background-color: {C_CARD_H};
                border-radius: 14px;
                border-left: 4px solid {self._color};
            }}
        """)
        super().enterEvent(event)

    def leaveEvent(self, event):
        self.setStyleSheet(f"""
            JobCard {{
                background-color: {C_CARD};
                border-radius: 14px;
                border-left: 4px solid {self._color};
            }}
        """)
        super().leaveEvent(event)


# ── Stats Panel ───────────────────────────────────────────────────────────────
class StatsPanel(QFrame):
    def __init__(self, jobs: list, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""
            StatsPanel {{
                background-color: {C_SIDEBAR};
                border-bottom: 1px solid {C_DIVIDER};
            }}
        """)
        self.setFixedHeight(62)
        self._build(jobs)

    def _build(self, jobs: list):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(20, 8, 20, 8)
        layout.setSpacing(0)

        total = len(jobs)
        if total == 0:
            return

        with_contact = sum(
            1 for j in jobs if j.get("contact_email") or j.get("contact_phone")
        )
        facility_ok  = sum(1 for j in jobs if j.get("facility_match", True))
        dists        = [j["distance_km"] for j in jobs if j.get("distance_km") is not None]
        avg_dist     = (sum(dists) / len(dists)) if dists else None
        high_score   = sum(1 for j in jobs if (j.get("match_score") or 0) >= 70)

        def _stat(val: str, caption: str, col: str = C_TEXT):
            w = QWidget()
            w.setStyleSheet("background: transparent;")
            vl = QVBoxLayout(w)
            vl.setContentsMargins(14, 0, 14, 0)
            vl.setSpacing(1)
            v = QLabel(val)
            v.setStyleSheet(f"color: {col}; font-size: 17px; font-weight: 800;")
            v.setAlignment(Qt.AlignmentFlag.AlignCenter)
            c = QLabel(caption)
            c.setStyleSheet(f"color: {C_SUB}; font-size: 9px; letter-spacing: 0.3px;")
            c.setAlignment(Qt.AlignmentFlag.AlignCenter)
            vl.addWidget(v)
            vl.addWidget(c)
            return w

        def _divider():
            d = QFrame()
            d.setFrameShape(QFrame.Shape.VLine)
            d.setFixedWidth(1)
            d.setStyleSheet(f"background: {C_DIVIDER}; border: none; margin: 6px 0;")
            return d

        layout.addWidget(_stat(str(total), "Stellen gesamt", C_TEXT))
        layout.addWidget(_divider())
        pct_c = int(with_contact / total * 100)
        layout.addWidget(_stat(
            f"{pct_c}%",
            f"{with_contact} mit Kontakt",
            "#22C55E" if pct_c >= 60 else C_WARN_FG,
        ))
        layout.addWidget(_divider())
        pct_f = int(facility_ok / total * 100)
        layout.addWidget(_stat(
            f"{pct_f}%", "Einrichtung passend",
            C_PRIMARY if pct_f >= 60 else C_WARN_FG,
        ))
        layout.addWidget(_divider())
        if avg_dist is not None:
            layout.addWidget(_stat(
                f"Ø {avg_dist:.0f} km", "Ø Entfernung",
                C_PRIMARY if avg_dist <= 30 else C_WARN_FG,
            ))
            layout.addWidget(_divider())
        layout.addWidget(_stat(
            str(high_score), "Score ≥ 70%",
            "#22C55E" if high_score > 0 else C_SUB,
        ))
        layout.addStretch()


# ── Sidebar ───────────────────────────────────────────────────────────────────
class Sidebar(QWidget):
    search_clicked = pyqtSignal(str, str, str, str, str, str, int)
    pdf_loaded     = pyqtSignal(str)
    save_profile   = pyqtSignal()
    load_profile   = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedWidth(320)
        self.setStyleSheet(f"background-color: {C_SIDEBAR};")
        self._history_items: list = []
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Logo ──────────────────────────────────────────────────────────
        # Sidebar ist 320px breit – Logo soll die gesamte Breite ausfüllen.
        header_bar = QWidget()
        header_bar.setFixedHeight(260)
        header_bar.setStyleSheet(f"background-color: {C_SIDEBAR};")
        hl = QHBoxLayout(header_bar)
        hl.setContentsMargins(0, 0, 0, 0)
        hl.setSpacing(0)
        if os.path.exists(LOGO_PATH):
            lbl = QLabel()
            # Quadratisches Logo (2048×2048) auf Sidebar-Breite hochskalieren
            lbl.setPixmap(QPixmap(LOGO_PATH).scaled(
                320, 260, Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation,
            ))
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setStyleSheet("background: transparent;")
            hl.addWidget(lbl)
        else:
            lbl = QLabel("MEDthief")
            lbl.setStyleSheet(f"color: {C_PRIMARY}; font-size: 48px; font-weight: bold;")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            hl.addWidget(lbl)
        root.addWidget(header_bar)

        # ── Profil-Leiste ─────────────────────────────────────────────────
        prof_bar = QWidget()
        prof_bar.setStyleSheet(
            f"background-color: {C_CARD}; border-bottom: 1px solid {C_DIVIDER};"
        )
        pl = QHBoxLayout(prof_bar)
        pl.setContentsMargins(12, 6, 12, 6)
        pl.setSpacing(6)

        self.profile_combo = QComboBox()
        self.profile_combo.setEditable(True)
        self.profile_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.profile_combo.lineEdit().setPlaceholderText("Profilname …")
        self.profile_combo.setStyleSheet(f"""
            QComboBox {{
                background-color: #141A22; color: {C_TEXT};
                border: 1px solid #2F3A48; border-radius: 8px;
                padding: 5px 10px; font-size: 11px;
            }}
            QComboBox QLineEdit {{
                color: {C_TEXT}; background: transparent;
                font-size: 11px; border: none; padding: 0;
            }}
            QComboBox QLineEdit::placeholder {{ color: #94A3B8; }}
            QComboBox:focus {{ border-color: {C_PRIMARY}; }}
            QComboBox::drop-down {{ border: none; width: 20px; }}
            QComboBox::down-arrow {{
                border-left: 3px solid transparent;
                border-right: 3px solid transparent;
                border-top: 4px solid {C_SUB};
                width: 0; height: 0; margin-right: 8px;
            }}
            QComboBox QAbstractItemView {{
                background-color: #141A22; color: {C_TEXT};
                border: 1px solid #2F3A48; selection-background-color: {C_PRIMARY};
            }}
        """)
        self._refresh_profile_list()
        pl.addWidget(self.profile_combo, stretch=1)

        for icon, tip, sig in [
            ("💾", "Profil speichern", self.save_profile),
            ("📂", "Profil laden",    None),
        ]:
            btn = QPushButton(icon)
            btn.setToolTip(tip)
            btn.setFixedSize(28, 28)
            if sig:
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background-color: #22{C_PRIMARY[1:]}; color: {C_PRIMARY_L};
                        border: 1px solid #55{C_PRIMARY[1:]}; border-radius: 6px;
                        font-size: 13px; padding: 0;
                    }}
                    QPushButton:hover {{ background-color: #44{C_PRIMARY[1:]}; }}
                """)
                btn.clicked.connect(sig)
            else:
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {C_CARD_H}; color: {C_SUB};
                        border: 1px solid {C_BORDER}; border-radius: 6px;
                        font-size: 13px; padding: 0;
                    }}
                    QPushButton:hover {{ color: {C_TEXT}; }}
                """)
                btn.clicked.connect(self._on_load_profile)
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            pl.addWidget(btn)

        root.addWidget(prof_bar)

        # ── Scrollbarer innerer Bereich ───────────────────────────────────
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.Shape.NoFrame)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        # Viewport-Margins erzwingen Abstand innerhalb des ScrollArea-Viewports
        scroll_area.setViewportMargins(12, 10, 12, 10)
        scroll_area.setStyleSheet(f"""
            QScrollArea {{ background: {C_SIDEBAR}; border: none; }}
            QScrollBar:vertical {{
                background: transparent; width: 4px; margin: 0;
                border: none;
            }}
            QScrollBar::handle:vertical {{
                background: #2A4540; border-radius: 2px; min-height: 20px;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
        """)
        scroll_area.viewport().setStyleSheet(f"background: {C_SIDEBAR};")

        inner = QWidget()
        inner.setStyleSheet(f"background-color: {C_SIDEBAR};")
        layout = QVBoxLayout(inner)
        layout.setContentsMargins(0, 0, 0, 4)
        layout.setSpacing(0)
        scroll_area.setWidget(inner)
        root.addWidget(scroll_area, stretch=1)

        # ── PDF ───────────────────────────────────────────────────────────
        self._section(layout, "LEBENSLAUF")
        self.btn_pdf = QPushButton("📄  Lebenslauf laden")
        self.btn_pdf.setStyleSheet(f"""
            QPushButton {{
                background-color: #22{C_PRIMARY[1:]}; color: {C_PRIMARY_L};
                border: 1.5px solid #55{C_PRIMARY[1:]}; border-radius: 8px;
                padding: 8px 12px; font-size: 12px; font-weight: 600;
            }}
            QPushButton:hover {{
                background-color: #40{C_PRIMARY[1:]}; border-color: {C_PRIMARY}; color: white;
            }}
        """)
        self.btn_pdf.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_pdf.clicked.connect(self._pick_pdf)
        layout.addWidget(self.btn_pdf)
        layout.addSpacing(3)

        self.lbl_pdf = QLabel("Kein PDF geladen")
        self.lbl_pdf.setStyleSheet("color: #AEDBD5; font-size: 10px;")
        self.lbl_pdf.setWordWrap(True)
        layout.addWidget(self.lbl_pdf)

        # ── Suchbegriff ───────────────────────────────────────────────────
        self._section(layout, "SUCHBEGRIFF")
        from cv_parser import MEDWING_ALL_TITLES
        self.entry_job = QComboBox()
        self.entry_job.setEditable(True)
        self.entry_job.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        # Wichtig: Mindestbreite NICHT nach längstem Eintrag berechnen,
        # sonst zwingt MEDWING_ALL_TITLES das Layout auf 400+ px!
        self.entry_job.setSizeAdjustPolicy(
            QComboBox.SizeAdjustPolicy.AdjustToMinimumContentsLengthWithIcon
        )
        self.entry_job.setMinimumContentsLength(10)
        self.entry_job.addItem("")
        self.entry_job.addItems(MEDWING_ALL_TITLES)
        self.entry_job.setCurrentIndex(0)
        self.entry_job.setStyleSheet(self._field_style())
        self.entry_job.lineEdit().setPlaceholderText("z. B. Pflegefachkraft")
        self._set_placeholder_color(self.entry_job.lineEdit())
        c = self.entry_job.completer()
        if c:
            c.setFilterMode(Qt.MatchFlag.MatchContains)
            c.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        layout.addWidget(self.entry_job)

        # ── Adresse ───────────────────────────────────────────────────────
        self._section(layout, "ADRESSE")
        self.entry_addr = self._make_lineedit("z. B. Hamburg")
        layout.addWidget(self.entry_addr)

        # ── Einrichtungsart ───────────────────────────────────────────────
        self._section(layout, "EINRICHTUNGSART")
        self.entry_einrichtung = self._make_lineedit("z. B. Krankenhaus")
        layout.addWidget(self.entry_einrichtung)

        # ── Fachabteilung ─────────────────────────────────────────────────
        self._section(layout, "FACHABTEILUNG")
        from cv_parser import FACHABTEILUNGEN as _FA_LIST
        self.entry_dept = MultiSelectButton(_FA_LIST, "Fachabteilung wählen …")
        layout.addWidget(self.entry_dept)

        # ── Arbeitszeit ───────────────────────────────────────────────────
        self._section(layout, "ARBEITSZEIT")
        self.combo_arbeitszeit = self._make_combo(
            ["Egal", "Vollzeit", "Teilzeit", "Vollzeit / Teilzeit"]
        )
        layout.addWidget(self.combo_arbeitszeit)

        # ── Schicht ───────────────────────────────────────────────────────
        self._section(layout, "SCHICHT")
        self.combo_schicht = self._make_combo(
            ["Egal", "Tagdienst", "Früh/Spät", "Wechselschicht", "Dauernacht"]
        )
        layout.addWidget(self.combo_schicht)

        # ── Radius ────────────────────────────────────────────────────────
        self._section(layout, "RADIUS")
        r_row = QHBoxLayout()
        r_row.setSpacing(8)
        self.slider = QSlider(Qt.Orientation.Horizontal)
        self.slider.setRange(5, 100)
        self.slider.setValue(25)
        self.lbl_radius = QLabel("25 km")
        self.lbl_radius.setFixedWidth(44)
        self.lbl_radius.setStyleSheet(
            f"color: {C_TEXT}; font-size: 12px; font-weight: 700;"
        )
        self.slider.valueChanged.connect(lambda v: self.lbl_radius.setText(f"{v} km"))
        r_row.addWidget(self.slider)
        r_row.addWidget(self.lbl_radius)
        layout.addLayout(r_row)
        layout.addSpacing(16)

        # ── Such-Button ───────────────────────────────────────────────────
        self.btn_search = QPushButton("🔍  Jobs suchen")
        self.btn_search.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                    stop:0 #45C9B4, stop:1 #2A9880);
                color: white; border-radius: 9px;
                font-size: 13px; font-weight: 700; padding: 10px 18px;
                border: none;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                    stop:0 #52D9C4, stop:1 #38A694);
            }}
            QPushButton:pressed {{
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                    stop:0 #2A9880, stop:1 #1E7A65);
            }}
            QPushButton:disabled {{ background: {C_BORDER}; color: {C_SUB}; }}
        """)
        self.btn_search.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_search.clicked.connect(self._on_search)
        layout.addWidget(self.btn_search)
        layout.addSpacing(6)

        self.progress = QProgressBar()
        self.progress.setRange(0, 0)
        self.progress.hide()
        layout.addWidget(self.progress)
        layout.addSpacing(8)

        # ── Notizen ───────────────────────────────────────────────────────
        self._section(layout, "NOTIZEN")
        self.notes_edit = QTextEdit()
        self.notes_edit.setPlaceholderText("Notizen zum Kandidaten …")
        self.notes_edit.setFixedHeight(56)
        self.notes_edit.setStyleSheet(f"""
            QTextEdit {{
                background-color: #1D3430; color: {C_TEXT};
                border: 1px solid #2F3A48; border-radius: 8px;
                padding: 5px 9px; font-size: 11px;
            }}
            QTextEdit:focus {{ border: 1px solid {C_PRIMARY}; }}
        """)
        # Placeholder-Farbe via Palette
        _np = self.notes_edit.palette()
        _np.setColor(QPalette.ColorRole.PlaceholderText, QColor("#AEDBD5"))
        self.notes_edit.setPalette(_np)
        layout.addWidget(self.notes_edit)
        layout.addSpacing(8)

        # ── Zuletzt gesucht ───────────────────────────────────────────────
        self._section(layout, "ZULETZT GESUCHT")
        self._history_container = QWidget()
        self._history_container.setStyleSheet("background: transparent;")
        self._history_layout = QVBoxLayout(self._history_container)
        self._history_layout.setContentsMargins(0, 0, 0, 0)
        self._history_layout.setSpacing(3)
        layout.addWidget(self._history_container)
        layout.addStretch()

        self.lbl_status = QLabel("Bereit.")
        self.lbl_status.setStyleSheet("color: #AEDBD5; font-size: 10px;")
        self.lbl_status.setWordWrap(True)
        layout.addWidget(self.lbl_status)

    # ── Helpers ───────────────────────────────────────────────────────────
    @staticmethod
    def _field_style() -> str:
        """CSS-String für QComboBox (Suchbegriff-Feld)."""
        return f"""
            QComboBox {{
                background-color: #1D3430;
                color: #F4F6F8;
                border: 1px solid #2F3A48;
                border-radius: 8px;
                padding: 7px 12px;
                font-size: 12px;
            }}
            QComboBox:focus {{ border: 1.5px solid #38A694; background-color: #223C38; }}
            QComboBox::drop-down {{ border: none; width: 24px; }}
            QComboBox::down-arrow {{
                image: none;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 5px solid #94A3B8;
                width: 0; height: 0; margin-right: 8px;
            }}
            QComboBox QAbstractItemView {{
                background-color: #141A22; color: #F4F6F8;
                border: 1px solid #2F3A48;
                selection-background-color: #38A694;
                selection-color: white; padding: 4px; outline: none;
            }}
        """

    @staticmethod
    def _set_placeholder_color(widget):
        """Setzt Placeholder-Farbe via QPalette (zuverlässiger als CSS in PyQt6)."""
        pal = widget.palette()
        pal.setColor(QPalette.ColorRole.PlaceholderText, QColor("#AEDBD5"))
        widget.setPalette(pal)

    @staticmethod
    def _make_combo(items: list) -> QComboBox:
        """Erstellt eine gut lesbare QComboBox."""
        cb = QComboBox()
        cb.setSizeAdjustPolicy(
            QComboBox.SizeAdjustPolicy.AdjustToMinimumContentsLengthWithIcon
        )
        cb.setMinimumContentsLength(8)
        cb.addItems(items)
        cb.setStyleSheet(f"""
            QComboBox {{
                background-color: #1D3430;
                color: #F4F6F8;
                border: 1px solid #2F3A48;
                border-radius: 10px;
                padding: 9px 14px;
                font-size: 13px;
            }}
            QComboBox:focus {{ border: 1.5px solid #38A694; background-color: #223C38; }}
            QComboBox::drop-down {{ border: none; width: 28px; }}
            QComboBox::down-arrow {{
                image: none;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 5px solid #94A3B8;
                width: 0; height: 0; margin-right: 10px;
            }}
            QComboBox QAbstractItemView {{
                background-color: #141A22; color: #F4F6F8;
                border: 1px solid #2F3A48;
                selection-background-color: #38A694;
                selection-color: white; padding: 4px; outline: none;
            }}
        """)
        return cb

    @staticmethod
    def _make_lineedit(placeholder: str) -> QLineEdit:
        """Erstellt ein QLineEdit mit gut sichtbarem Placeholder via Palette."""
        le = QLineEdit()
        le.setPlaceholderText(placeholder)
        le.setStyleSheet(f"""
            QLineEdit {{
                background-color: #1D3430;
                color: #F4F6F8;
                border: 1px solid #2F3A48;
                border-radius: 10px;
                padding: 9px 14px;
                font-size: 13px;
            }}
            QLineEdit:focus {{
                border: 1.5px solid #38A694;
                background-color: #223C38;
            }}
        """)
        # Placeholder-Farbe via QPalette setzen (zuverlässiger als ::placeholder in PyQt6)
        pal = le.palette()
        pal.setColor(QPalette.ColorRole.PlaceholderText, QColor("#AEDBD5"))
        le.setPalette(pal)
        return le

    def _section(self, layout, text):
        layout.addSpacing(16)
        row = QHBoxLayout()
        row.setSpacing(7)
        accent = QFrame()
        accent.setFixedSize(3, 11)
        accent.setStyleSheet(f"background: {C_PRIMARY}; border-radius: 1px;")
        lbl = QLabel(text)
        lbl.setStyleSheet(f"""
            color: {C_MUTED}; font-size: 10px;
            font-weight: 600; letter-spacing: 1.2px;
        """)
        row.addWidget(accent)
        row.addSpacing(4)
        row.addWidget(lbl)
        row.addStretch()
        layout.addLayout(row)
        layout.addSpacing(6)

    def _pick_pdf(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Lebenslauf auswählen", "",
            "PDF Dateien (*.pdf);;Alle Dateien (*.*)"
        )
        if path:
            self.pdf_loaded.emit(path)

    def _on_search(self):
        job  = self.entry_job.currentText().strip()
        addr = self.entry_addr.text().strip()
        if not job and not addr:
            self.set_status("Bitte Suchbegriff oder Adresse eingeben.")
            return
        einrichtung = self.entry_einrichtung.text().strip()
        dept        = self.entry_dept.text_value().strip()
        arbeitszeit = self.combo_arbeitszeit.currentText()
        schicht     = self.combo_schicht.currentText()
        r           = self.slider.value()
        if job or addr:
            entry = (job, addr, einrichtung, dept, arbeitszeit, schicht)
            self._history_items = [e for e in self._history_items if e != entry]
            self._history_items.insert(0, entry)
            self._history_items = self._history_items[:5]
            self._rebuild_history()
        self.search_clicked.emit(job, addr, einrichtung, dept, arbeitszeit, schicht, r)

    def _rebuild_history(self):
        while self._history_layout.count():
            item = self._history_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        if not self._history_items:
            l = QLabel("Noch keine Suchen")
            l.setStyleSheet(f"color: #A8D0CB; font-size: 10px; font-style: italic;")
            self._history_layout.addWidget(l)
            return
        for entry in self._history_items:
            job, addr, *_ = entry
            city = re.sub(r"^\d{4,5}\s*", "", addr.split(",")[0].strip()).strip() \
                   or addr.split(",")[0].strip()
            parts = [p for p in [job, city] if p]
            chip  = " · ".join(parts) or "Unbekannt"
            if len(chip) > 35:
                chip = chip[:34] + "…"
            btn = QPushButton(chip)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: #1D3430; color: #A8D0CB;
                    border: 1px solid #2F3A48; border-radius: 6px;
                    font-size: 10px; padding: 4px 10px; text-align: left;
                }}
                QPushButton:hover {{ background-color: #243E3A; color: {C_TEXT}; }}
            """)
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            btn.clicked.connect(
                lambda _, e=entry: self.set_fields(e[0], e[1], e[2], e[3], e[4], e[5])
            )
            self._history_layout.addWidget(btn)

    def _on_load_profile(self):
        name = self.profile_combo.currentText().strip()
        if name:
            self.load_profile.emit(name)

    def _refresh_profile_list(self):
        current = self.profile_combo.currentText() \
            if hasattr(self, "profile_combo") else ""
        self.profile_combo.clear()
        for p in ProfileManager.list_profiles():
            self.profile_combo.addItem(p)
        if current:
            self.profile_combo.setCurrentText(current)

    # ── Public API ────────────────────────────────────────────────────────
    def set_pdf_label(self, name: str):
        self.lbl_pdf.setText(f"✅  {name}")
        self.lbl_pdf.setStyleSheet(f"color: {C_PRIMARY_L}; font-size: 11px; font-weight: 600;")

    def set_fields(self, job_title="", address="", einrichtung="",
                   fachabteilung="", arbeitszeit="", schicht=""):
        if job_title:
            idx = self.entry_job.findText(job_title)
            self.entry_job.setCurrentIndex(idx) if idx >= 0 \
                else self.entry_job.setCurrentText(job_title)
        if address:     self.entry_addr.setText(address)
        if einrichtung: self.entry_einrichtung.setText(einrichtung)
        if fachabteilung: self.entry_dept.set_from_text(fachabteilung)
        if arbeitszeit:
            idx = self.combo_arbeitszeit.findText(arbeitszeit)
            if idx >= 0: self.combo_arbeitszeit.setCurrentIndex(idx)
        if schicht:
            idx = self.combo_schicht.findText(schicht)
            if idx >= 0: self.combo_schicht.setCurrentIndex(idx)

    def set_status(self, msg: str):
        self.lbl_status.setText(msg)

    def set_searching(self, active: bool):
        self.btn_search.setEnabled(not active)
        self.btn_search.setText("Wird gesucht …" if active else "🔍  Jobs suchen")
        self.progress.show() if active else self.progress.hide()

    def get_profile_name(self) -> str:
        return self.profile_combo.currentText().strip()

    def set_profile_name(self, name: str):
        self.profile_combo.setCurrentText(name)


# ── Ergebnisbereich ───────────────────────────────────────────────────────────
class ResultsPanel(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"background-color: {C_BG};")
        self._cards         = []
        self._all_jobs      : list = []
        self._shown_count   = 0
        self._load_more_btn = None
        self._stats_panel   = None
        self._ctx           : dict = {}
        self._build()

    def _build(self):
        self._root_layout = QVBoxLayout(self)
        self._root_layout.setContentsMargins(0, 0, 0, 0)
        self._root_layout.setSpacing(0)

        # ── Header ────────────────────────────────────────────────────────
        header = QWidget()
        header.setStyleSheet(f"""
            background-color: {C_SIDEBAR};
            border-bottom: 1px solid {C_DIVIDER};
        """)
        h_outer = QVBoxLayout(header)
        h_outer.setContentsMargins(20, 8, 20, 8)
        h_outer.setSpacing(6)

        # Row 1: count label + Excel button
        h1 = QHBoxLayout()
        self.lbl_count = QLabel("Suchergebnisse")
        self.lbl_count.setStyleSheet(
            f"color: {C_TEXT}; font-size: 15px; font-weight: 700;"
        )
        h1.addWidget(self.lbl_count)
        h1.addStretch()

        if _OPENPYXL_OK:
            self.btn_excel = QPushButton("📊 Excel")
            self.btn_excel.setStyleSheet(f"""
                QPushButton {{
                    background-color: {C_PRIMARY_D}; color: white;
                    font-size: 11px; font-weight: bold;
                    padding: 5px 12px; border-radius: 6px;
                    border: 1px solid {C_PRIMARY};
                }}
                QPushButton:hover {{ background-color: {C_PRIMARY}; }}
            """)
            self.btn_excel.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            self.btn_excel.clicked.connect(self._export_excel)
            self.btn_excel.hide()
            h1.addWidget(self.btn_excel)
        else:
            self.btn_excel = None

        h_outer.addLayout(h1)

        # Row 2: Source filter buttons
        h2 = QHBoxLayout()
        h2.setSpacing(4)
        self._source_btns = {}
        self._active_source = None  # None = alle

        btn_all = QPushButton("Alle")
        btn_all.setCheckable(True)
        btn_all.setChecked(True)
        btn_all.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_all.setStyleSheet(f"""
            QPushButton {{
                color: {C_TEXT}; font-size: 10px; font-weight: 700;
                padding: 4px 10px; border-radius: 10px;
                background-color: transparent; border: 1px solid {C_BORDER};
            }}
            QPushButton:hover {{
                background-color: rgba(255,255,255,0.08);
            }}
            QPushButton:checked {{
                background-color: {C_PRIMARY}; color: #0A0A0B;
                border: 1px solid {C_PRIMARY}; font-weight: 800;
            }}
        """)
        btn_all.clicked.connect(lambda: self._filter_source(None))
        h2.addWidget(btn_all)
        self._source_btns["_all"] = btn_all

        for src, col in JobCard.SOURCE_COLORS.items():
            btn = QPushButton(JobCard.SOURCE_DISPLAY.get(src, src))
            btn.setCheckable(True)
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            btn.setStyleSheet(f"""
                QPushButton {{
                    color: {C_SUB}; font-size: 10px; font-weight: 600;
                    padding: 4px 10px; border-radius: 10px;
                    background-color: transparent;
                    border: 1px solid {C_BORDER};
                }}
                QPushButton:hover {{
                    color: {col}; background-color: #18{col[1:]};
                    border: 1px solid #66{col[1:]};
                }}
                QPushButton:checked {{
                    color: #0A0A0B; background-color: {col};
                    border: 1px solid {col}; font-weight: 800;
                }}
            """)
            btn.clicked.connect(lambda checked, s=src: self._filter_source(s))
            h2.addWidget(btn)
            self._source_btns[src] = btn

        h2.addStretch()
        h_outer.addLayout(h2)

        self._root_layout.addWidget(header)

        # ── Scroll-Bereich ────────────────────────────────────────────────
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll.setStyleSheet(f"background-color: {C_BG}; border: none;")

        self.inner = QWidget()
        self.inner.setStyleSheet(f"background-color: {C_BG};")
        self.inner_layout = QVBoxLayout(self.inner)
        self.inner_layout.setContentsMargins(20, 16, 20, 16)
        self.inner_layout.setSpacing(8)

        self.lbl_placeholder = QLabel(
            "Lade einen Lebenslauf und klicke auf 'Jobs suchen'."
        )
        self.lbl_placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_placeholder.setStyleSheet(f"color: {C_SUB}; font-size: 14px;")
        # Placeholder sitzt im Layout, nimmt aber keinen Platz wenn versteckt
        sp = self.lbl_placeholder.sizePolicy()
        sp.setRetainSizeWhenHidden(False)
        self.lbl_placeholder.setSizePolicy(sp)
        self.inner_layout.addWidget(
            self.lbl_placeholder, alignment=Qt.AlignmentFlag.AlignCenter
        )
        # Bottom-Stretch, damit Cards oben beginnen und nicht vertikal aufgeblasen werden
        self.inner_layout.addStretch(1)

        self.scroll.setWidget(self.inner)
        self._root_layout.addWidget(self.scroll)

    # ─────────────────────────────────────────────────────────────────────
    def clear(self):
        for card in self._cards:
            card.setParent(None)
            card.deleteLater()
        self._cards.clear()
        if self._load_more_btn:
            self._load_more_btn.setParent(None)
            self._load_more_btn.deleteLater()
            self._load_more_btn = None
        if self._stats_panel:
            self._stats_panel.setParent(None)
            self._stats_panel.deleteLater()
            self._stats_panel = None
        self.lbl_placeholder.hide()
        if self.btn_excel:
            self.btn_excel.hide()

    def show_jobs(self, jobs: list, ctx: dict):
        self.clear()
        self._ctx = ctx
        if not jobs:
            self.lbl_placeholder.setText(
                "Keine passenden Stellen gefunden.\n"
                "Versuche einen größeren Radius oder andere Einrichtungsart."
            )
            self.lbl_placeholder.show()
            self.lbl_count.setText("Keine Ergebnisse")
            return

        self._all_jobs = jobs
        self._shown_count = 0
        self.lbl_count.setText(f"{len(jobs)} Stellen gefunden")
        if self.btn_excel:
            self.btn_excel.show()

        self._show_source_counts(jobs)

        # StatsPanel nach dem Header einsetzen
        self._stats_panel = StatsPanel(jobs)
        self._root_layout.insertWidget(1, self._stats_panel)

        self._render_more(20)

        # Seen-URLs aktualisieren
        seen = ctx.get("seen_urls", set())
        for j in jobs:
            u = j.get("url", "")
            if u:
                seen.add(u)

    def _show_source_counts(self, jobs):
        counts: dict = {}
        for j in jobs:
            s = j.get("source", "")
            counts[s] = counts.get(s, 0) + 1
        # Button-Texte aktualisieren und leere verstecken
        for src, btn in self._source_btns.items():
            if src == "_all":
                btn.setText(f"Alle ({len(jobs)})")
                btn.show()
            else:
                n = counts.get(src, 0)
                if n > 0:
                    btn.setText(f"{JobCard.SOURCE_DISPLAY.get(src, src)} ({n})")
                    btn.show()
                else:
                    btn.hide()

    def _filter_source(self, source):
        """Filter-Jobs nach Quelle. source=None zeigt alle."""
        self._active_source = source
        # Buttons updaten
        for src, btn in self._source_btns.items():
            if source is None:
                btn.setChecked(src == "_all")
            else:
                btn.setChecked(src == source)
        # Cards filtern
        for card in self._cards:
            job_src = card.job.get("source", "")
            if source is None or job_src == source:
                card.show()
            else:
                card.hide()
        # Zähler updaten
        visible = sum(1 for c in self._cards if not c.isHidden())
        total = len(self._all_jobs)
        if source:
            self.lbl_count.setText(f"{visible}/{total} Stellen ({JobCard.SOURCE_DISPLAY.get(source, source)})")
        else:
            self.lbl_count.setText(f"{total} Stellen gefunden")

    def _stretch_index(self) -> int:
        """Index des bottom-stretch items im inner_layout."""
        return self.inner_layout.count() - 1

    def _render_more(self, batch: int = 20):
        if self._load_more_btn:
            self._load_more_btn.setParent(None)
            self._load_more_btn.deleteLater()
            self._load_more_btn = None

        end = min(self._shown_count + batch, len(self._all_jobs))
        for job in self._all_jobs[self._shown_count:end]:
            card = JobCard(job, self._ctx, self.inner)
            # Respect active source filter
            if self._active_source and job.get("source", "") != self._active_source:
                card.hide()
            # Vor dem Bottom-Stretch einfügen
            self.inner_layout.insertWidget(self._stretch_index(), card)
            self._cards.append(card)
        self._shown_count = end

        remaining = len(self._all_jobs) - self._shown_count
        if remaining > 0:
            btn = QPushButton(f"Mehr laden ({remaining} weitere)")
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: transparent; color: {C_SUB};
                    border: 1px solid {C_BORDER}; border-radius: 10px;
                    padding: 10px 24px; font-size: 12px; font-weight: 500;
                }}
                QPushButton:hover {{ color: {C_TEXT}; border-color: {C_PRIMARY}; }}
            """)
            btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            btn.clicked.connect(lambda: self._render_more(20))
            self.inner_layout.insertWidget(
                self._stretch_index(), btn,
                alignment=Qt.AlignmentFlag.AlignCenter
            )
            self._load_more_btn = btn

    def show_placeholder(self, text: str):
        self.clear()
        self.lbl_placeholder.setText(text)
        self.lbl_placeholder.show()

    def _export_excel(self):
        if not _OPENPYXL_OK or not self._all_jobs:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel speichern", "jobs_export.xlsx", "Excel Dateien (*.xlsx)"
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Jobs"
            headers = [
                "Titel", "Firma", "Ort", "Entfernung (km)", "Quelle",
                "Ansprechpartner", "Telefon", "E-Mail", "Email-Status",
                "Einrichtungstyp", "Match-Score", "CRM-Status", "URL",
            ]
            ws.append(headers)
            hfill = PatternFill("solid", fgColor="1E534B")
            for cell in ws[1]:
                cell.font = XLFont(bold=True, color="FFFFFF")
                cell.fill = hfill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.freeze_panes = "A2"
            job_status = self._ctx.get("job_status", {})
            for job in self._all_jobs:
                dist = job.get("distance_km")
                ws.append([
                    job.get("title", ""),
                    job.get("company", ""),
                    job.get("location", ""),
                    f"{dist:.1f}" if dist is not None else "",
                    job.get("source", ""),
                    job.get("contact_name", ""),
                    job.get("contact_phone", ""),
                    job.get("contact_email", ""),
                    "⚠ geschätzt" if job.get("contact_email_guessed")
                    else ("✓" if job.get("contact_email") else ""),
                    job.get("detected_facility", ""),
                    f"{job['match_score']}%" if job.get("match_score") is not None else "",
                    job_status.get(job.get("url", ""), "Offen"),
                    job.get("url", ""),
                ])
            for i, w in enumerate(
                [40, 30, 25, 14, 14, 30, 20, 35, 14, 25, 12, 14, 60], 1
            ):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            wb.save(path)
            QMessageBox.information(self, "Export erfolgreich", f"Gespeichert:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export-Fehler", str(e))


# ── Kandidaten-Tab ────────────────────────────────────────────────────────────
class CandidateTab(QWidget):
    """Eigenständiger Arbeitsbereich pro Kandidat."""

    # Signal: Parsed CV-Ergebnis zurück an Main-Thread liefern (thread-safe)
    _cv_parsed = pyqtSignal(dict)

    @staticmethod
    def _last_name(full_name: str) -> str:
        """Extrahiert den Nachnamen aus einem vollen Namen.
        Berücksichtigt deutsche Präfixe wie 'von', 'van', 'de', 'zu'."""
        parts = [p for p in (full_name or "").strip().split() if p]
        if not parts:
            return ""
        if len(parts) == 1:
            return parts[0]
        # "Max von Müller" → "von Müller"; "Anna van der Berg" → "van der Berg"
        prefixes = {"von", "van", "de", "der", "den", "du", "zu", "zur", "le", "la"}
        for i, p in enumerate(parts):
            if p.lower() in prefixes:
                return " ".join(parts[i:])
        return parts[-1]

    def __init__(self, tab_label: str = "Neuer Kandidat", parent=None):
        super().__init__(parent)
        self._tab_label    = tab_label
        self._profile_name = ""
        self._job_status   : dict = {}
        self._job_notes    : dict = {}
        self._seen_urls    : set  = set()
        self._search_params : dict = {}
        self._candidate_info: dict = {}
        self._thread       = None
        self._worker       = None
        self.cv_parser     = CVParser()
        self.job_searcher  = JobSearcher()
        self._cv_parsed.connect(self._apply_cv)
        self._build()

    def _build(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self.sidebar = Sidebar()
        self.sidebar.pdf_loaded.connect(self._load_pdf)
        self.sidebar.search_clicked.connect(self._start_search)
        self.sidebar.save_profile.connect(self._save_profile)
        self.sidebar.load_profile.connect(self._load_profile_by_name)
        layout.addWidget(self.sidebar)

        self.results = ResultsPanel()
        layout.addWidget(self.results, stretch=1)

    # ── CV ────────────────────────────────────────────────────────────────
    def _load_pdf(self, path: str):
        self.sidebar.set_pdf_label(path.split("/")[-1])
        self.sidebar.set_status("PDF wird analysiert …")

        def _run():
            try:
                result = self.cv_parser.parse(path)
            except Exception as e:
                import traceback
                traceback.print_exc()
                result = {"error": str(e)}
            # Signal emittieren → thread-sicherer Aufruf im Main-Thread
            self._cv_parsed.emit(result)

        threading.Thread(target=_run, daemon=True).start()

    def _apply_cv(self, result: dict):
        if result.get("error"):
            self.sidebar.set_status(f"Fehler: {result['error']}")
            return

        def _c(v, n=60):
            return (v or "").split("\n")[0].strip()[:n]

        job_title   = _c(result.get("job_title", ""), 60)
        einrichtung = _c(result.get("facility_type", ""), 40)
        dept        = _c(result.get("fachabteilungen", ""), 50)
        raw_addr    = result.get("wohnort", "") or result.get("location", "")
        wohnort     = re.sub(r",?\s*Deutschland.*", "", raw_addr, flags=re.I).strip()
        wohnort     = _c(wohnort, 80)

        stelle_raw = (result.get("stelle_typ", "") or "").lower()
        if "vollzeit" in stelle_raw and "teilzeit" in stelle_raw:
            arbeitszeit = "Vollzeit / Teilzeit"
        elif "vollzeit" in stelle_raw:
            arbeitszeit = "Vollzeit"
        elif "teilzeit" in stelle_raw:
            arbeitszeit = "Teilzeit"
        else:
            arbeitszeit = ""

        sch_raw = (result.get("schichten", "") or "").lower()
        if "dauernacht" in sch_raw or ("nacht" in sch_raw and "tag" not in sch_raw):
            schicht = "Dauernacht"
        elif "wechsel" in sch_raw:
            schicht = "Wechselschicht"
        elif "tag" in sch_raw and "früh" not in sch_raw and "spät" not in sch_raw:
            schicht = "Tagdienst"
        elif "früh" in sch_raw or "spät" in sch_raw:
            schicht = "Früh- & Spätschicht"
        else:
            schicht = ""

        self.sidebar.set_fields(job_title, wohnort, einrichtung, dept, arbeitszeit, schicht)

        # Kandidaten-Infos für Akquise-Email speichern
        self._candidate_info = {
            "name":           _c(result.get("name", ""), 40),
            "job_title":      job_title,
            "einrichtung":    einrichtung,
            "fachabteilungen": dept,
            "verfuegbar_ab":  _c(result.get("verfuegbar_ab", ""), 30),
            "wohnort":        wohnort,
            "arbeitszeit":    arbeitszeit,
            "schichten":      schicht,
        }

        # Tab-Bezeichnung → Nachname des Kandidaten
        cname = _c(result.get("name", ""), 40)
        if cname:
            self._tab_label = self._last_name(cname) or cname
            self._refresh_tab_text()

        lines = []
        if result.get("is_medwing"):
            lines.append("✅ MEDWING Kurzprofil erkannt")
        for k, v in [
            ("Beruf", job_title), ("Adresse", wohnort[:40]),
            ("Einrichtung", einrichtung), ("Abt.", dept[:50]),
            ("Arbeitszeit", arbeitszeit), ("Schicht", schicht),
        ]:
            if v:
                lines.append(f"{k}: {v}")
        if result.get("verfuegbar_ab"):
            lines.append(f"Verfügbar: {result['verfuegbar_ab']}")
        self.sidebar.set_status("\n".join(lines) or "Felder bitte prüfen.")

        if job_title or wohnort:
            r = self.sidebar.slider.value()
            QTimer.singleShot(300, lambda: self._start_search(
                job_title, wohnort, einrichtung, dept, arbeitszeit, schicht, r
            ))

    # ── Suche ─────────────────────────────────────────────────────────────
    def _start_search(self, job_title, address, einrichtung,
                      department, arbeitszeit="", schicht="", radius=25):
        self._search_params = dict(
            job_title=job_title, einrichtung=einrichtung, radius=radius
        )
        self.sidebar.set_searching(True)
        self.results.show_placeholder("Suche läuft …")

        self._thread = QThread()
        self._worker = SearchWorker(
            self.job_searcher, job_title, address, department,
            einrichtung, radius, arbeitszeit, schicht,
        )
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.progress.connect(self.sidebar.set_status)
        self._worker.finished.connect(self._on_results)
        self._worker.error.connect(lambda e: self.sidebar.set_status(f"Fehler: {e}"))
        self._worker.finished.connect(self._thread.quit)
        self._worker.error.connect(self._thread.quit)
        self._thread.finished.connect(lambda: self.sidebar.set_searching(False))
        self._thread.start()

    def _on_results(self, jobs: list):
        from job_searcher import _is_relevant
        p = self._search_params
        search_title = p.get("job_title", "")

        # Post-Filter: fachfremde Stellen entfernen (z.B. Zahnmedizin bei Pflege-Suche)
        if search_title:
            jobs = [j for j in jobs if _is_relevant(j.get("title", ""), search_title)]

        for job in jobs:
            job["match_score"] = compute_match_score(
                job, search_title,
                p.get("einrichtung", ""), p.get("radius", 30),
            )
        # Pflegia zuerst, dann nach Facility-Match, Score, Entfernung
        _SOURCE_PRIO = {"pflegia": 0}
        jobs_sorted = sorted(
            jobs,
            key=lambda j: (
                0 if j.get("facility_match", True) else 1,
                _SOURCE_PRIO.get(j.get("source", ""), 5),
                -(j.get("match_score") or 0),
                j.get("distance_km") or 9999,
            ),
        )
        ctx = {
            "job_status":       self._job_status,
            "job_notes":        self._job_notes,
            "seen_urls":        self._seen_urls,
            "on_status_change": self._on_status_change,
            "on_notes_change":  self._on_notes_change,
            "candidate_info":   self._candidate_info,
        }
        self.results.show_jobs(jobs_sorted, ctx)

        if self._profile_name:
            ProfileManager.save_job_data(
                self._profile_name, self._job_status,
                self._job_notes, self._seen_urls,
            )

        _mac_notify(
            "MEDthief",
            f"{len(jobs_sorted)} Stellen für {p.get('job_title', 'Suche')} gefunden",
        )

    def _on_status_change(self, url: str, status: str):
        self._job_status[url] = status
        if self._profile_name:
            ProfileManager.save_job_data(
                self._profile_name, self._job_status,
                self._job_notes, self._seen_urls,
            )

    def _on_notes_change(self, url: str, note: str):
        self._job_notes[url] = note
        if self._profile_name:
            ProfileManager.save_job_data(
                self._profile_name, self._job_status,
                self._job_notes, self._seen_urls,
            )

    # ── Profil speichern ──────────────────────────────────────────────────
    def _save_profile(self):
        name = self.sidebar.get_profile_name()
        if not name:
            name, ok = QInputDialog.getText(
                self, "Profil speichern", "Profilname (z. B. Kandidatenname):",
            )
            if not ok or not name.strip():
                return
            name = name.strip()
            self.sidebar.set_profile_name(name)
        self._profile_name = name
        data = ProfileManager.build_from_sidebar(self.sidebar, name)
        data["job_status"]     = self._job_status
        data["job_notes"]      = self._job_notes
        data["seen_urls"]      = list(self._seen_urls)
        data["candidate_name"] = self._tab_label or name
        ProfileManager.save(name, data)
        self.sidebar._refresh_profile_list()
        self.sidebar.set_status(f"✅ Profil gespeichert: {name}")

    # ── Profil laden ──────────────────────────────────────────────────────
    def _load_profile_by_name(self, name: str):
        profile = ProfileManager.load(name)
        if not profile:
            self.sidebar.set_status(f"Profil '{name}' nicht gefunden.")
            return
        self._profile_name = name
        self._job_status, self._job_notes, self._seen_urls = \
            ProfileManager.load_job_data(name)
        ProfileManager.apply_to_sidebar(profile, self.sidebar)
        self.sidebar.set_profile_name(name)
        cname = profile.get("candidate_name", "")
        if cname:
            self._tab_label = self._last_name(cname) or cname
            self._refresh_tab_text()
        self.sidebar.set_status(f"✅ Profil geladen: {name}")

    def get_tab_label(self) -> str:
        return self._tab_label

    def _refresh_tab_text(self):
        """Aktualisiert den Tab-Titel im umschließenden QTabWidget sofort."""
        w = self.parent()
        while w is not None and not isinstance(w, QTabWidget):
            w = w.parent()
        if w is None:
            return
        idx = w.indexOf(self)
        if idx >= 0:
            w.setTabText(idx, self._tab_label)


# ── Hauptfenster ──────────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("MEDthief")
        self.resize(1360, 840)
        self.setMinimumSize(960, 620)

        central = QWidget()
        self.setCentralWidget(central)
        ml = QVBoxLayout(central)
        ml.setContentsMargins(0, 0, 0, 0)
        ml.setSpacing(0)

        self.tabs = QTabWidget()
        self.tabs.setTabsClosable(True)
        self.tabs.tabCloseRequested.connect(self._close_tab)
        self.tabs.currentChanged.connect(self._sync_tab_title)

        # "+" Button rechts oben im Tab-Bar
        btn_plus = QPushButton("＋")
        btn_plus.setToolTip("Neuen Kandidaten-Tab öffnen")
        btn_plus.setStyleSheet(f"""
            QPushButton {{
                background: transparent; color: {C_SUB};
                border: none; font-size: 18px; font-weight: bold;
                padding: 2px 12px; margin-right: 4px;
            }}
            QPushButton:hover {{ color: {C_TEXT}; }}
        """)
        btn_plus.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_plus.clicked.connect(self._add_tab)
        self.tabs.setCornerWidget(btn_plus, Qt.Corner.TopRightCorner)

        ml.addWidget(self.tabs)
        self._add_tab()

    def _add_tab(self):
        n = self.tabs.count() + 1
        tab = CandidateTab(f"Kandidat {n}")
        idx = self.tabs.addTab(tab, f"Kandidat {n}")
        self.tabs.setCurrentIndex(idx)

    def _close_tab(self, index: int):
        if self.tabs.count() <= 1:
            # Letzten Tab zurücksetzen statt schließen
            tab = self.tabs.widget(0)
            if isinstance(tab, CandidateTab):
                tab.results.show_placeholder(
                    "Lade einen Lebenslauf und klicke auf 'Jobs suchen'."
                )
                tab.sidebar.set_status("Bereit.")
        else:
            self.tabs.removeTab(index)

    def _sync_tab_title(self, index: int):
        tab = self.tabs.widget(index)
        if isinstance(tab, CandidateTab):
            self.tabs.setTabText(index, tab.get_tab_label())


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyleSheet(STYLESHEET)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
